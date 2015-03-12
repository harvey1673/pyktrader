'''
Created on Apr 10, 2014

@author: Harvey
'''
import pulp
import openpyxl

def optionmargin(exchange, fwd, strikes, prem, otype):
    if otype == 'c':
        vv = [max(k-fwd,0) for k in strikes];
    else:
        vv = [max(fwd-k,0) for k in strikes];
       
    if exchange == 'CFFE':
        marginRatio = 0.15;
        minRatio = 0.667;        
        margin=[p+max(marginRatio*fwd-v, minRatio*marginRatio*fwd) for (p, v) in zip(prem, vv)];
    
    return margin;        

def strat2pos(init_strat):    
    uc = init_strat['callsprd'];
    up = init_strat['putsprd'];
    vc = init_strat['callfly'];
    vp = init_strat['putfly'];
    wf = init_strat['callput'];
    nSize = len(wf);
    call = [0.0]*nSize;
    put  = [0.0]*nSize;
    call[0] = uc[0] + vc[1] + wf[0];
    call[nSize-1] =-uc[nSize-2] + vc[nSize-2] + wf[nSize-1];
    put[0]  =-up[0] + vp[1] - wf[0];
    put[nSize-1] = up[nSize-2] + vp[nSize-2] - wf[nSize-1];
    
    for i in range(1, nSize-1):
        call[i] = uc[i]-uc[i-1] + wf[i] + vc[i-1]-2*vc[i]+vc[i+1];
        put[i]  =-up[i]+up[i-1] - wf[i] + vp[i-1]-2*vp[i]+vp[i+1];          
    return {'call':call, 'put':put}

def arboptimizer( mktdata, exchange, init_strat, capital_limit, fwdmargin ):
    fwdBid = mktdata['fwdbid'];
    fwdAsk = mktdata['fwdask'];
    fwdMid = (fwdBid + fwdAsk)/2.0;
    strikes = mktdata['strike'];
    callBids = mktdata['callbid'];
    callAsks = mktdata['callask'];
    callMids = [ (a + b)/2 for (a,b) in zip(callBids, callAsks)];
    putBids = mktdata['putbid'];
    putAsks = mktdata['putask'];
    putMids = [ (a + b)/2 for (a,b) in zip(putBids, putAsks)];
    nSize = len(strikes);
    
    uc0 = init_strat['callsprd'];
    up0 = init_strat['putsprd'];
    vc0 = init_strat['callfly'];
    vp0 = init_strat['putfly'];
    wf0 = init_strat['callput'];  

    margin_c = optionmargin(exchange, fwdMid, strikes, callMids, 'c');
    margin_p = optionmargin(exchange, fwdMid, strikes, putMids, 'p');
    prob = pulp.LpProblem("ArbAllocator",pulp.LpMaximize);

    uc = pulp.LpVariable.dict('uc_%s', range(nSize-1));
    up = pulp.LpVariable.dict('up_%s', range(nSize-1));
    vc = pulp.LpVariable.dict('vc_%s', range(nSize));
    vp = pulp.LpVariable.dict('vp_%s', range(nSize));
    wf = pulp.LpVariable.dict('wf_%s', range(nSize));    
    xuc = pulp.LpVariable.dict('xuc_%s', range(nSize-1), lowBound=0.0);
    xup = pulp.LpVariable.dict('xup_%s', range(nSize-1), lowBound=0.0);   
    yc = pulp.LpVariable.dict('yc_%s', range(nSize), lowBound=0.0);
    yp = pulp.LpVariable.dict('yp_%s', range(nSize), lowBound=0.0);
    yf = pulp.LpVariable('yf', lowBound = 0.0, cat='Continuous');    
    zc = pulp.LpVariable.dict('zc_%s', range(nSize), lowBound=0.0);
    zp = pulp.LpVariable.dict('zp_%s', range(nSize), lowBound=0.0);
    zf = pulp.LpVariable('zf', lowBound = 0.0, cat='Continuous');
        
    for i in vc.viewkeys():
        vc[i].lowBound = -vc0[i];
         
    for i in vp.viewkeys():
        vp[i].lowBound = -vp0[i];

    prob += sum([(uc0[i]+uc[i])*(callMids[i+1]-callMids[i]+strikes[i+1]-strikes[i]) \
                -(strikes[i+1]-strikes[i])*xuc[i] for i in range(nSize-1)]) \
            + sum([(up0[i]+up[i])*(putMids[i]-putMids[i+1]+strikes[i+1]-strikes[i]) \
                -(strikes[i+1]-strikes[i])*xup[i] for i in range(nSize-1)]) \
            + sum([(vc0[i]+vc[i])*(2*callMids[i]-callMids[i-1]-callMids[i+1]) for i in range(1, nSize-1)]) \
            + sum([(vp0[i]+vp[i])*(2*putMids[i] - putMids[i-1] - putMids[i+1]) for i in range(1, nSize-1)]) \
            + sum([(wf0[i]+wf[i])*(putMids[i]-callMids[i]+fwdMid-strikes[i]) for i in range(nSize)]) \
            - sum([yc[i]*(callAsks[i]-callMids[i]) for i in range(nSize)]) \
            - sum([yp[i]*(putAsks[i]-putMids[i]) for i in range(nSize)]) \
            - yf*(fwdAsk-fwdMid);       
        
    prob += sum([zc[i]*margin_c[i]+zp[i]*margin_p[i] for i in range(nSize)])+ \
        sum([(uc0[i]+uc[i])*(callMids[i]-callMids[i+1]) for i in range(nSize-1)]) + \
        sum([(up0[i]+up[i])*(putMids[i+1]-putMids[i]) for i in range(nSize-1)]) + \
        sum([(vc0[i]+vc[i])*(callMids[i-1]+callMids[i+1]-2*callMids[i]) for i in range(1, nSize-1)]) + \
        sum([(vp0[i]+vp[i])*(putMids[i-1] +putMids[i+1] -2*putMids[i] ) for i in range(1, nSize-1)]) + \
        sum([(wf0[i]+wf[i])*(callMids[i] - putMids[i]) for i in range(nSize)]) + \
        + zf*fwdmargin*fwdMid <=capital_limit;
    
    prob += zf - sum([wf[i]+wf0[i] for i in range(nSize)]) >= 0;
    prob += zf + sum([wf[i]+wf0[i] for i in range(nSize)]) >= 0;
    prob += yf - sum([wf[i] for i in range(nSize)]) >= 0;
    prob += yf + sum([wf[i] for i in range(nSize)]) >= 0;
    
    prob +=  (uc0[0]+uc[0])+ (vc0[1]+vc[1]) + (wf0[0]+wf[0]) + zc[0] >= 0;
    prob += -(up0[0]+up[0])+ (vp0[1]+vp[1]) - (wf0[0]+wf[0]) + zp[0] >= 0;
    prob += -uc0[nSize-2]+vc0[nSize-2]+wf0[nSize-1] - uc[nSize-2]+vc[nSize-2]+wf[nSize-1] + zc[nSize-1] >= 0;
    prob += up0[nSize-2]+vp0[nSize-2]-wf0[nSize-1] + up[nSize-2]+vp[nSize-2]-wf[nSize-1] + zp[nSize-1] >= 0;
    
    prob += yc[0]+uc[0]+vc[1]+wf[0] >= 0;
    prob += yc[0]-uc[0]-vc[1]-wf[0] >= 0;
    prob += yp[0]-up[0]+vp[1]-wf[0] >= 0;
    prob += yp[0]+up[0]-vp[1]+wf[0] >= 0;
    prob += yc[nSize-1]-uc[nSize-2]+vc[nSize-2]+wf[nSize-1] >= 0;
    prob += yc[nSize-1]+uc[nSize-2]-vc[nSize-2]-wf[nSize-1] >= 0;
    prob += yp[nSize-1]+up[nSize-2]+vp[nSize-2]-wf[nSize-1] >= 0;
    prob += yp[nSize-1]-up[nSize-2]-vp[nSize-2]+wf[nSize-1] >= 0;
    prob += vc[0] == 0.0;
    prob += vc[nSize-1] == 0.0;
    prob += vp[0] == 0.0;
    prob += vp[nSize-1] == 0.0;
    
    for i in range(nSize-1):
        prob += xuc[i] - uc0[i] - uc[i] >= 0
        prob += xup[i] - up0[i] - up[i] >= 0
        
    for i in range(1, nSize-1):        
        prob += uc0[i]-uc0[i-1]+vc0[i-1]+vc0[i+1]-2*vc0[i]+wf0[i]+uc[i]-uc[i-1]+vc[i-1]+vc[i+1]-2*vc[i]+wf[i]+zc[i] >= 0; 
        prob += up0[i-1]-up0[i]+vp0[i-1]+vp0[i+1]-2*vp0[i]-wf0[i]+up[i-1]-up[i]+vp[i-1]+vp[i+1]-2*vp[i]-wf[i]+zp[i] >= 0;
        prob += yc[i]+uc[i]-uc[i-1]+vc[i-1]+vc[i+1]-2*vc[i]+wf[i] >= 0;
        prob += yc[i]-uc[i]+uc[i-1]-vc[i-1]-vc[i+1]+2*vc[i]-wf[i] >= 0;
        prob += yp[i]+up[i-1]-up[i]+vp[i-1]+vp[i+1]-2*vp[i]-wf[i] >= 0;
        prob += yp[i]-up[i-1]+up[i]-vp[i-1]-vp[i+1]+2*vp[i]+wf[i] >= 0;

    prob.solve();    
    solutions = {};
    for v in prob.variables():
        solutions[v.name] = v.varValue
          
    callsprd=[ solutions['uc_'+str(i)] for i in range(nSize-1)] + [0];
    putsprd =[ solutions['up_'+str(i)] for i in range(nSize-1)] + [0];
    callput= [ solutions['wf_'+str(i)] for i in range(nSize)];        
    callfly =[ solutions['vc_'+str(i)] for i in range(nSize) ];
    putfly = [ solutions['vp_'+str(i)] for i in range(nSize) ];     
    res = {'callput': callput, 
           'callsprd':callsprd, 
           'putsprd': putsprd, 
           'callfly': callfly, 
           'putfly' : putfly, 
           'value' : prob.objective.value(), 
           'status': pulp.LpStatus[prob.status]};
    return res;
        
def arboptimizer2( mktdata, exchange, capital_limit, fwdmargin ):
    fwdBid = mktdata['fwdbid'];
    fwdAsk = mktdata['fwdask'];
    fwdMid = (fwdBid + fwdAsk)/2.0;
    strikes = mktdata['strike'];
    callBids = mktdata['callbid'];
    callAsks = mktdata['callask'];
    callMids = [ (a + b)/2 for (a,b) in zip(callBids, callAsks)];
    putBids = mktdata['putbid'];
    putAsks = mktdata['putask'];
    putMids = [ (a + b)/2 for (a,b) in zip(putBids, putAsks)];
    nSize = len(strikes);

    margin_c = optionmargin(exchange, fwdMid, strikes, callMids, 'c');
    margin_p = optionmargin(exchange, fwdMid, strikes, putMids, 'p');
    prob = pulp.LpProblem("ArbAllocator",pulp.LpMaximize);

    wc = pulp.LpVariable.dict('wc_%s', range(nSize-1), lowBound=0.0);
    wp = pulp.LpVariable.dict('wp_%s', range(nSize-1), lowBound=0.0);
    uc = pulp.LpVariable.dict('uc_%s', range(nSize-1), lowBound=0.0);
    up = pulp.LpVariable.dict('up_%s', range(nSize-1), lowBound=0.0);
    vc = pulp.LpVariable.dict('vc_%s', range(nSize), lowBound=0.0);
    vp = pulp.LpVariable.dict('vp_%s', range(nSize), lowBound=0.0);
    rc = pulp.LpVariable.dict('rc_%s', range(nSize), lowBound=0.0);
    rp = pulp.LpVariable.dict('rp_%s', range(nSize), lowBound=0.0);
    zc = pulp.LpVariable.dict('zc_%s', range(nSize), lowBound=0.0);
    zp = pulp.LpVariable.dict('zp_%s', range(nSize), lowBound=0.0);
    zf = pulp.LpVariable('zf', lowBound = 0.0, cat='Continuous');

    prob += sum([wc[i]*(callBids[i+1]-callAsks[i]) for i in range(nSize-1)]) + \
        sum([wp[i]*(putBids[i]-putAsks[i+1]) for i in range(nSize-1)]) + \
        sum([uc[i]*(callBids[i]-callAsks[i+1]-strikes[i+1]+strikes[i]) for i in range(nSize-1)]) + \
        sum([up[i]*(putBids[i+1]-putAsks[i]-strikes[i+1]+strikes[i]) for i in range(nSize-1)]) + \
        sum([vc[i]*(2*callBids[i] - callAsks[i-1] - callAsks[i+1]) for i in range(1, nSize-1)]) + \
        sum([vp[i]*(2*putBids[i] - putAsks[i-1] - putAsks[i+1]) for i in range(1, nSize-1)]) + \
        sum([rc[i]*(putBids[i] - callAsks[i] + fwdBid - strikes[i]) for i in range(nSize)]) + \
        sum([rp[i]*(callBids[i] - putAsks[i] - fwdAsk + strikes[i]) for i in range(nSize)])
        
    prob += sum([zc[i]*margin_c[i]+zp[i]*margin_p[i] for i in range(nSize)])+ \
        sum([(wc[i]-uc[i])*(callMids[i]-callMids[i+1]) for i in range(nSize-1)]) + \
        sum([(wp[i]-up[i])*(putMids[i+1]-putMids[i]) for i in range(nSize-1)]) + \
        sum([vc[i]*(callMids[i-1]+callMids[i+1]-2*callMids[i]) for i in range(1, nSize-1)]) + \
        sum([vp[i]*(putMids[i-1] +putMids[i+1] -2*putMids[i] ) for i in range(1, nSize-1)]) + \
        sum([(rc[i]-rp[i])*(callMids[i] - putMids[i]) for i in range(nSize)]) + \
        + zf*fwdmargin*fwdMid <=capital_limit;
    prob += sum([rp[i]-rc[i] for i in range(nSize)]) - zf <= 0;
    prob += sum([rc[i]-rp[i] for i in range(nSize)]) - zf <= 0;
    prob += wc[0]-uc[0]+vc[1]+rc[0]-rp[0] + zc[0] >= 0;
    prob += -wp[0]+up[0]+vp[1]+rp[0]-rc[0]+ zp[0] >= 0;
    prob += -wc[nSize-2]+uc[nSize-2]+vc[nSize-2]+rc[nSize-1]-rp[nSize-1] + zc[nSize-1] >= 0;
    prob += wp[nSize-2]-up[nSize-2]+vp[nSize-2]+rp[nSize-1]-rc[nSize-1]+ zp[nSize-1] >= 0;
    prob += vc[0] == 0.0;
    prob += vc[nSize-1] == 0.0;
    prob += vp[0] == 0.0;
    prob += vp[nSize-1] == 0.0;
    for i in range(1, nSize-1):
        prob += wc[i]-uc[i]-(wc[i-1]-uc[i-1])+vc[i-1]+vc[i+1]-2*vc[i]+rc[i]-rp[i]+zc[i] >= 0; 
        prob += wp[i-1]-up[i-1]-(wp[i]-up[i])+vp[i-1]+vp[i+1]-2*vp[i]+rp[i]-rc[i]+zp[i] >= 0;

    prob.solve();    
    solutions = {};
    for v in prob.variables():
        solutions[v.name] = v.varValue
          
    callsprd=[ solutions['wc_'+str(i)]-solutions['uc_'+str(i)] for i in range(nSize-1)];
    putsprd =[ solutions['wp_'+str(i)]-solutions['up_'+str(i)] for i in range(nSize-1)];
    callput=[ solutions['rc_'+str(i)]-solutions['rp_'+str(i)] for i in range(nSize)];        
    callfly = [ solutions['vc_'+str(i)] for i in range(1,nSize-1) ];
    putfly = [ solutions['vp_'+str(i)] for i in range(1, nSize-1) ];     
    res = {'callput': callput, 
           'callsprd':callsprd, 
           'putsprd': putsprd, 
           'callfly': callfly, 
           'putfly' : putfly, 
           'value' : prob.objective.value(), 
           'status': pulp.LpStatus[prob.status]};
    return res;

if __name__ == "__main__":
    dataFile = 'test.xlsx';
    sheetName = 'mktData';
    wb = openpyxl.load_workbook('test.xlsx', data_only=True);
    ws = wb[sheetName];
    
    mktRange = 'A6:E18';
    wr = ws.range(mktRange);
    mktdata = {};
    mktdata['fwdbid'] = float(ws.cell(row=2,column=0).value);
    mktdata['fwdask'] = float(ws.cell(row=2,column=1).value);      
    for i in range(len(wr)):
        row = wr[i];
        for j in range(len(row)):
            colname = str(wr[0][j].value).lower();
            if i == 0:
                mktdata[colname] = [];
            else:
                mktdata[colname].append(float(wr[i][j].value));            

#     init_strat = {'callsprd':[0]*8,
#                 'putsprd':[0]*8,
#                 'callfly':[0]*8,
#                 'putfly':[0]*8,
#                 'callput':[0,0,0,0,0,-1500,0,0],
#                 }
    init_strat = {};
    posRange = 'G6:K18';
    wr = ws.range(posRange);  
    for i in range(len(wr)):
        row = wr[i];
        for j in range(len(row)):
            colname = str(wr[0][j].value).lower();
            if i == 0:
                init_strat[colname] = [];
            else:
                init_strat[colname].append(float(wr[i][j].value));
                    
    exchange = 'CFFE';
    fwdmargin = 0.12;
    capital_limit = 1000000;

    res1 = arboptimizer( mktdata, exchange, init_strat, capital_limit, fwdmargin );
    #res2 = arboptimizer2( mktdata, exchange, capital_limit, fwdmargin );
    pos = strat2pos(res1);
    res1['call'] = pos['call'];
    res1['put']  = pos['put'];
    outfile = "test_output.xlsx";
    outRange = 'N6:T18';
    wr = ws.range(outRange); 
    for i, col in enumerate(['callsprd','putsprd','callfly','putfly','callput','call','put']):
        wr[0][i].value = col;
        for j, pos in enumerate(res1[col]):
            wr[j+1][i].value = pos;  
    
    wb.save(outfile);   
    pass
    