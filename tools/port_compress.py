import ms.version
import csv
ms.version.addpkg("numpy", "1.7.1-mkl")
ms.version.addpkg("pulp", "1.5.3")
ms.version.addpkg("jdcal", "1.0")
ms.version.addpkg("et_xmlfile", "1.0.1")
ms.version.addpkg("openpyxl", "2.4.5")
#ms.version.addpkg("ibm_db", "2.0.4-9.7.4")

import pulp
import openpyxl
#import ibm_db
import datetime 
import csv

def get_db_cashflow(sdate, edate, dbdate, tradelist, leg):
    return 0

def compress_trade():
    dataPath ='S:\\Harvey\\compression\\AUDUSD_Feb17\\'
    dataFile = 'input_audusd_xccs_170321.xlsx'
    outFile = 'result_audusd_xccs_170321a'
    
    wb = openpyxl.load_workbook(dataPath+dataFile, read_only = True)
    sheetName = 'Trend Static Data'
    ws = wb[sheetName]

    lower_w = -10.0
    upper_w = 10.0

    datefor = '%d-%b-%Y'
    tradeData = {'tickets':[], 'notionals':[], 'maturities':[], 'seasonalities':[]}
    lower_mat = datetime.datetime.now()+datetime.timedelta(days=0)
    upper_mat = datetime.datetime.now()+datetime.timedelta(days=70*365)
    
    pv_limit = 1000000.0
    ccy_pv_limit = {'USD':(-1000000, 1000000), 'AUD':(-1000000, 1000000)}
    para_pv01_limit = 5000
    pv01_limit = 5000
    season_lower = 0
    season_upper = 20000
    
    use_season = True
    use_tenor_limits = False
    use_trade_fixes = False
    use_ccy_PV = True
    use_PV_bnd = False
    pv_lowerbnd = 1000000

    key_col = 0
    notl_col = 10
    season_col = 2
    mat_col = 19
    
    for idx , row in enumerate(ws.iter_rows()):
        if (idx == 0): 
            print row[0]
            for i, istr in enumerate(row):                
                if 'Ticket' in istr.value: key_col = i
                if 'Current Notional' in istr.value: notl_col = i
                if 'Termination Date' in istr.value: 
                    mat_col = i
                if use_season and 'Seasonality' in istr.value: season_col = i
        if (idx > 0):
            print row[mat_col].value
            mat = datetime.datetime.strptime(str(row[mat_col].value), datefor)
            if (mat > lower_mat ) and (mat < upper_mat):
                tradeData['tickets'].append(str(row[key_col].value))
                tradeData['notionals'].append(float(str(row[notl_col].value).replace(',','')))
                tradeData['maturities'].append(mat)
                if use_season:
                    tradeData['seasonalities'].append(float(str(row[season_col].value).replace(',','')))
    print len(tradeData['tickets'])
    sheetName = 'PV'
    ws = wb[sheetName]    
    pvmap = {}
    key_col = 3
    ccy_col = 1
    pv_col  = 10
    for idx , row in enumerate(ws.iter_rows()):
        if (idx == 0):
            for i, istr in enumerate(row):
                if 'Ticket' in istr.value: key_col = i
                if 'ALL' in istr.value: pv_col = i
                if 'Currency' in istr.value: ccy_col = i
        else:
            ccy = str(row[ccy_col].value)
            if ccy not in pvmap.keys():
                pvmap[ccy]= {}
            key = str(row[key_col].value)
            if (key in tradeData['tickets']):
                if key in pvmap[ccy].keys():
                    pvmap[ccy][key] += float(str(row[pv_col].value).replace(',',''))
                else:
                    pvmap[ccy][key] = float(str(row[pv_col].value).replace(',',''))     
    print "Loading trade data is done"
    sheetName = 'Risk'
    ws = wb[sheetName]
    riskmap = {}
    partialmap = {}
    all_processes = []
    all_tenors = []
    proc_col = 3
    key_col = 4
    partial_col = 11
    for idx , row in enumerate(ws.iter_rows()):
        if idx == 0:
            for i, istr in enumerate(row):
                if 'Ticket' in istr.value: key_col = i
                if 'Process' in istr.value: proc_col = i
                if 'Partial' in istr.value: partial_col = i
                if i>partial_col: all_tenors.append(istr.value)
        else:
            processkey = str(row[proc_col].value)
            if processkey not in all_processes:
                all_processes.append(processkey)
            
            ticketkey = str(row[key_col].value)
            ncol = len(row)
            if ticketkey in tradeData['tickets']:           
                key = tuple( [ticketkey, processkey] )
                if key in partialmap.keys():
                    partialmap[key] += float(str(row[partial_col].value).replace(',',''))
                else:
                    partialmap[key] = float(str(row[partial_col].value).replace(',',''))
            
                if key in riskmap.keys():
                    thisrow = riskmap[key]
                else:
                    thisrow = [ 0 for i in range(partial_col+1,ncol)]
            
                for i in range(partial_col+1, ncol):
                    if row[i].value != None:
                        thisrow[i-(partial_col+1)] += float(str(row[i].value).replace(',',''))
            
                riskmap[key] = thisrow
    print "Loading risk is done"
    trade_map = {}
    if use_trade_fixes:
        sheetName = 'TradeWeight'
        ws = wb[sheetName] 
        for idx , row in enumerate(ws.iter_rows()):
            if row[0] in tradeData['tickets']:
                trade_map[row[0]] = float(row[1])
                
    tenor_map = {}
    if use_tenor_limits:
        sheetName = 'TenorLimit'
        ws = wb[sheetName] 
        for idx , row in enumerate(ws.iter_rows()):
            ten = row[0].value
            if ten in all_tenors:
                tenor_map[ten] = float(str(row[1].value).replace(',',''))               
    print "Setting up model"        
    prob = pulp.LpProblem("TradeCompression",pulp.LpMinimize)    
    w = pulp.LpVariable.dict('w_%s', tradeData['tickets'], lowBound=lower_w, upBound=upper_w, cat='Continuous')        
    if use_trade_fixes:
        for t in trade_map.keys():
            prob += w[t] == trade_map[t]
        
    if lower_w < 0:
        v = pulp.LpVariable.dict('v_%s', tradeData['tickets'], lowBound=0, upBound=upper_w, cat='Continuous')
        for t in w.viewkeys():
            prob += w[t] - v[t] <= 0
            prob +=-w[t] - v[t] <= 0
        prob += sum([v[t]*notl for t, notl in zip(tradeData['tickets'], tradeData['notionals'])]) 
    else:
        prob += sum([w[t]*notl for t, notl in zip(tradeData['tickets'], tradeData['notionals'])]) 
    print "Setting up ccy PV constraint", pvmap.keys()     
    if use_ccy_PV:
        for ccy in pvmap.keys():
            prob += sum([(w[t]-1)*pvmap[ccy][t] for t in tradeData['tickets']]) <= ccy_pv_limit[ccy][1] 
            prob += sum([(1-w[t])*pvmap[ccy][t] for t in tradeData['tickets']]) <= -ccy_pv_limit[ccy][0]         
    else:
        print pvmap.keys()
        prob += sum([(w[t]-1)*pvmap[c][t] for t in tradeData['tickets'] for c in pvmap.keys()]) <= pv_limit 
        prob += sum([(1-w[t])*pvmap[c][t] for t in tradeData['tickets'] for c in pvmap.keys()]) <= pv_limit
    if use_PV_bnd:
        prob += sum([(w[t]-1)*pvmap[c][t] for t in tradeData['tickets'] for c in pvmap.keys()]) >= pv_lowerbnd
    print "Setting up risk constraint"  
    for proc in all_processes:
        prob += sum([(w[key[0]]-1)*partialmap[key] for key in partialmap if key[1]==proc]) <= para_pv01_limit
        prob += sum([(1-w[key[0]])*partialmap[key] for key in partialmap if key[1]==proc]) <= para_pv01_limit
    
    if use_tenor_limits:
        for i, ten in enumerate(all_tenors):
            if ten in tenor_map.keys():
                prob += sum([(w[key[0]]-1)*riskmap[key][i] for key in partialmap])  <= tenor_map[ten]
                prob += sum([(1-w[key[0]])*riskmap[key][i] for key in partialmap])  <= tenor_map[ten]   
    else:
        print ncol-partial_col
        for col in range(ncol-partial_col-1):
            prob += sum([(w[key[0]]-1)*riskmap[key][col] for key in partialmap])  <= pv01_limit
            prob += sum([(1-w[key[0]])*riskmap[key][col] for key in partialmap])  <= pv01_limit
    print "Setting up seasonality constraint" 
    if use_season:
        prob += sum([w[t]*s for t, s in zip(tradeData['tickets'], tradeData['seasonalities'])]) >= season_lower
        prob += sum([w[t]*s for t, s in zip(tradeData['tickets'], tradeData['seasonalities'])]) <= season_upper
    print "start solving"
    prob.solve()
    #print pulp.LpStatus[prob.status]
    weights = {}
    for v in prob.variables():
        if v.name[0]=='w':
            ticket = v.name[2:]
            weights[ticket] = v.varValue
    sln = prob.objective.value()
    
    diffs = [['','Before','After','Difference'],]
    row = ['notionals', sum(tradeData['notionals']), sum([ abs(weights[key])*notl for key, notl in zip(tradeData['tickets'],tradeData['notionals'])])]
    row.append(row[2]-row[1])
    diffs.append(row)
    for ccy in pvmap.keys():
        row = ['PV(%s)' % ccy, sum([pvmap[ccy][key] for key in tradeData['tickets']]), 
                    sum([pvmap[ccy][key]*weights[key] for key in tradeData['tickets']])]
        row.append(row[2]-row[1])
        diffs.append(row)
        
    for proc in all_processes:
        row = [proc,sum([partialmap[key] for key in partialmap if key[1]==proc]),
                    sum([weights[key[0]]*partialmap[key] for key in partialmap if key[1]==proc])]
        row.append(row[2]-row[1])
        diffs.append(row)
    
    for col in range(len(all_tenors)):
        row = [all_tenors[col], sum([riskmap[key][col] for key in partialmap]),
                    sum([weights[key[0]]*riskmap[key][col] for key in partialmap])]
        row.append(row[2]-row[1])
        diffs.append(row)

    if use_season:
        row = ['seasonalities', sum(tradeData['seasonalities']),
                sum([weights[t]*s for t, s in zip(tradeData['tickets'], tradeData['seasonalities'])])]
        row.append(row[2]-row[1])
        diffs.append(row)
    outfile1 = dataPath + outFile + "summary.csv"
    with open(outfile1, "wb") as f:
        writer = csv.writer(f)
        writer.writerows(diffs)
    rows = [[key, weights[key], notl, weights[key]*notl] for key, notl in zip(tradeData['tickets'],tradeData['notionals'])]         
    rows = [['tickets','weight','original','after']] + rows
    outfile2 = dataPath + outFile + "weights.csv"
    with open(outfile2, "wb") as f:
        writer = csv.writer(f)
        writer.writerows(rows)   
    return (sln, weights, diffs)
    
if __name__ == "__main__":   
    value = compress_trade()     
    pass
